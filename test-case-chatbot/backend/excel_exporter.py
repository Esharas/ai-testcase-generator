import os
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# --------------------------------------------------
# Location of your Excel template
# --------------------------------------------------

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "CR Name_Test Case Document.xlsx"
)


# --------------------------------------------------
# Copy formatting from one row to another
# --------------------------------------------------

def copy_row_format(
    worksheet,
    source_row,
    target_row
):

    # Copy height
    source_height = worksheet.row_dimensions[
        source_row
    ].height

    if source_height is not None:
        worksheet.row_dimensions[
            target_row
        ].height = source_height

    # Copy formatting for columns A:M
    for column in range(1, 14):

        source_cell = worksheet.cell(
            source_row,
            column
        )

        target_cell = worksheet.cell(
            target_row,
            column
        )

        if source_cell.has_style:

            target_cell._style = copy(
                source_cell._style
            )

        if source_cell.number_format:
            target_cell.number_format = (
                source_cell.number_format
            )

        if source_cell.alignment:
            target_cell.alignment = copy(
                source_cell.alignment
            )

        if source_cell.protection:
            target_cell.protection = copy(
                source_cell.protection
            )


# --------------------------------------------------
# Remove existing merged cells in Test Cases area
# --------------------------------------------------

def remove_test_case_merges(worksheet):

    merges_to_remove = []

    for merged_range in worksheet.merged_cells.ranges:

        # We only remove merges that are in
        # the Test Cases data area.

        if (
            merged_range.min_row >= 2
            and merged_range.max_row >= 2
            and merged_range.min_col <= 13
        ):

            merges_to_remove.append(
                str(merged_range)
            )

    for merged_range in merges_to_remove:

        worksheet.unmerge_cells(
            merged_range
        )


# --------------------------------------------------
# Export test cases
# --------------------------------------------------

def export_test_cases_to_excel(
    test_cases,
    output_path
):

    # --------------------------------------------------
    # Check template
    # --------------------------------------------------

    if not os.path.exists(TEMPLATE_PATH):

        raise FileNotFoundError(
            "Excel template not found: "
            + TEMPLATE_PATH
        )


    # --------------------------------------------------
    # Load original template
    # --------------------------------------------------

    workbook = load_workbook(
        TEMPLATE_PATH
    )


    # Make sure Test Cases sheet exists
    if "Test Cases" not in workbook.sheetnames:

        raise ValueError(
            "The Excel template does not contain "
            "'Test Cases' sheet."
        )


    worksheet = workbook["Test Cases"]


    # --------------------------------------------------
    # Remember the original formatting rows
    # --------------------------------------------------

    template_max_row = worksheet.max_row


    # --------------------------------------------------
    # Remove old merged cells
    # --------------------------------------------------

    remove_test_case_merges(
        worksheet
    )


    # --------------------------------------------------
    # Clear old test case values
    #
    # Keep the header row.
    # --------------------------------------------------

    for row in range(
        2,
        worksheet.max_row + 1
    ):

        for column in range(
            1,
            14
        ):

            cell = worksheet.cell(
                row,
                column
            )

            # MergedCell objects cannot be edited
            if isinstance(
                cell,
                MergedCell
            ):
                continue

            cell.value = None


    # --------------------------------------------------
    # Start writing test cases
    # --------------------------------------------------

    current_row = 2


    for test_case in test_cases:

        steps = test_case.get(
            "steps",
            []
        )


        # If Gemini somehow returns no steps
        if not steps:

            steps = [
                {
                    "step": "",
                    "test_data": "",
                    "expected_result": ""
                }
            ]


        # --------------------------------------------------
        # Number of rows required for this test case
        #
        # Minimum 5 rows to match the template structure
        # --------------------------------------------------

        number_of_rows = max(
            len(steps),
            5
        )


        start_row = current_row

        end_row = (
            start_row
            + number_of_rows
            - 1
        )


        # --------------------------------------------------
        # Make sure enough rows exist
        # --------------------------------------------------

        while worksheet.max_row < end_row:

            new_row = worksheet.max_row + 1

            # Copy formatting from template row 2
            copy_row_format(
                worksheet,
                2,
                new_row
            )


        # --------------------------------------------------
        # Copy template formatting to generated rows
        # --------------------------------------------------

        for i in range(number_of_rows):

            target_row = (
                start_row + i
            )

            # Use the first 5 template rows
            # as formatting references
            source_row = (
                2 + min(i, 4)
            )

            copy_row_format(
                worksheet,
                source_row,
                target_row
            )


        # --------------------------------------------------
        # Fill each row
        # --------------------------------------------------

        for i in range(number_of_rows):

            row = start_row + i


            # ----------------------------------------------
            # Get step information
            # ----------------------------------------------

            if i < len(steps):

                step_data = steps[i]

            else:

                step_data = {
                    "step": "",
                    "test_data": "",
                    "expected_result": ""
                }


            # ----------------------------------------------
            # Test Case ID
            # ----------------------------------------------

            worksheet.cell(
                row,
                1
            ).value = (
                test_case.get(
                    "test_case_id",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Requirement ID
            # ----------------------------------------------

            worksheet.cell(
                row,
                2
            ).value = (
                test_case.get(
                    "requirement_id",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Title
            # ----------------------------------------------

            worksheet.cell(
                row,
                3
            ).value = (
                test_case.get(
                    "title",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Description
            # ----------------------------------------------

            worksheet.cell(
                row,
                4
            ).value = (
                test_case.get(
                    "description",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Test Step
            # ----------------------------------------------

            worksheet.cell(
                row,
                5
            ).value = step_data.get(
                "step",
                ""
            )


            # ----------------------------------------------
            # Test Data
            # ----------------------------------------------

            worksheet.cell(
                row,
                6
            ).value = step_data.get(
                "test_data",
                ""
            )


            # ----------------------------------------------
            # Expected Results
            # ----------------------------------------------

            worksheet.cell(
                row,
                7
            ).value = step_data.get(
                "expected_result",
                ""
            )


            # ----------------------------------------------
            # Test Results
            # ----------------------------------------------

            worksheet.cell(
                row,
                8
            ).value = (
                test_case.get(
                    "test_results",
                    "Not Executed"
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Priority
            # ----------------------------------------------

            worksheet.cell(
                row,
                9
            ).value = (
                test_case.get(
                    "priority",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Severity
            # ----------------------------------------------

            worksheet.cell(
                row,
                10
            ).value = (
                test_case.get(
                    "severity",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Defect ID
            # ----------------------------------------------

            worksheet.cell(
                row,
                11
            ).value = (
                test_case.get(
                    "defect_id",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Assignee
            # ----------------------------------------------

            worksheet.cell(
                row,
                12
            ).value = (
                test_case.get(
                    "assignee",
                    ""
                )
                if i == 0
                else None
            )


            # ----------------------------------------------
            # Comment
            # ----------------------------------------------

            worksheet.cell(
                row,
                13
            ).value = (
                test_case.get(
                    "comment",
                    ""
                )
                if i == 0
                else None
            )


        # --------------------------------------------------
        # Merge Test Results column
        #
        # Example:
        #
        # 5 steps:
        # H2:H6
        #
        # Next test case:
        # H7:H11
        # --------------------------------------------------

        if number_of_rows > 1:

            worksheet.merge_cells(
                start_row=start_row,
                start_column=8,
                end_row=end_row,
                end_column=8
            )


        # --------------------------------------------------
        # Move to next test case
        # --------------------------------------------------

        current_row = (
            end_row + 1
        )


    # --------------------------------------------------
    # Final data row
    # --------------------------------------------------

    last_data_row = (
        current_row - 1
    )


    # --------------------------------------------------
    # Add filter
    # --------------------------------------------------

    if last_data_row >= 2:

        worksheet.auto_filter.ref = (
            f"A1:M{last_data_row}"
        )


    # --------------------------------------------------
    # Freeze header
    # --------------------------------------------------

    worksheet.freeze_panes = "A2"


    # --------------------------------------------------
    # Save final Excel file
    # --------------------------------------------------

    workbook.save(
        output_path
    )


    return output_path